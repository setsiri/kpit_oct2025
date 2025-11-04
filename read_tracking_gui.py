import os
import re
import threading
import queue
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


def _norm_str(x):
    if x is None:
        return None
    if isinstance(x, str):
        return x.strip()
    return str(x).strip()

def _cell_value(ws: Worksheet, coord: str):
    try:
        v = ws[coord].value
        return _norm_str(v)
    except Exception:
        return None

def Check_Yes_All_applicable_TC_Presented_And_No_Inapplicable_TC_Presented(ws: Worksheet, SP: str, SOP: str):
    """
    คืนค่าเป็น tuple: (Yes_All_applicable_TC_Presented, No_Inapplicable_TC_Presented)
    หรือ return 0 ถ้า row1/row2 หา matching ไม่เจอ

    ปรับตาม README:
    - อ่านค่าที่ไม่ว่างทั้งหมดในแถว 1 เพื่อหา SP แบบยืดหยุ่น
      กลุ่ม A = คอลัมน์ A..P, กลุ่ม Q = คอลัมน์ Q..X, กลุ่ม Y = คอลัมน์ Y.. ไปจนกว่าจะเจอช่องว่างแรก (สำหรับแถว 2)
    - เมื่อทราบกลุ่มแล้ว ให้หา SOP ในแถว 2 ภายในช่วงของกลุ่มนั้น
    """
    sheet_name = ws.title
    SP_norm = _norm_str(SP)
    SOP_norm = _norm_str(SOP)

    print(f"\n[DEBUG] ===== Sheet: {sheet_name} =====")
    print(f"[DEBUG] SP from path = {SP_norm}")
    print(f"[DEBUG] SOP from path = {SOP_norm}")

    # อ่านทั้งแถว 1 เพื่อหาค่าที่ไม่ว่าง พร้อมพิมพ์ debug
    non_blank_row1 = []  # list[(col_idx, col_letter, value)]
    max_c = ws.max_column or 1
    for ci in range(1, max_c + 1):
        v = _norm_str(ws.cell(row=1, column=ci).value)
        if v is not None and v != "":
            col_letter = get_column_letter(ci)
            non_blank_row1.append((ci, col_letter, v))
    if non_blank_row1:
        dbg = ", ".join([f"{c}{1}: {v}" for (_, c, v) in non_blank_row1])
        print(f"[DEBUG] Row1 non-blanks -> {dbg}")
    else:
        print("[DEBUG] Row1 has no non-blank cells")
        return 0

    # เลือกกลุ่มจากตำแหน่ง cell ที่ค่าตรงกับ SP
    chosen_group = None  # 'A' | 'Q' | 'Y'
    chosen_group_example_cell = None
    if SP_norm:
        for ci, col_letter, v in non_blank_row1:
            if v == SP_norm:
                if 1 <= ci <= 16:  # A..P
                    chosen_group = 'A'
                elif 17 <= ci <= 24:  # Q..X
                    chosen_group = 'Q'
                else:  # Y..
                    chosen_group = 'Y'
                chosen_group_example_cell = f"{col_letter}1"
                break

    if chosen_group is None:
        print("[DEBUG] row1 can't find matching SP in any cell")
        return 0

    print(f"[DEBUG] Chosen group = {chosen_group} (matched at {chosen_group_example_cell})")

    # ค้นหา SOP ในแถว 2 ตามช่วงของกลุ่ม
    chosen_col = None  # column letter
    if chosen_group == 'A':
        start_ci, end_ci = 1, 16  # A..P
        for ci in range(start_ci, min(end_ci, max_c) + 1):
            val = _norm_str(ws.cell(row=2, column=ci).value)
            col_letter = get_column_letter(ci)
            print(f"[DEBUG] Checking {col_letter}2 = {val}")
            if SOP_norm and val and val == SOP_norm:
                chosen_col = col_letter
                break
    elif chosen_group == 'Q':
        start_ci, end_ci = 17, 24  # Q..X
        for ci in range(start_ci, min(end_ci, max_c) + 1):
            val = _norm_str(ws.cell(row=2, column=ci).value)
            col_letter = get_column_letter(ci)
            print(f"[DEBUG] Checking {col_letter}2 = {val}")
            if SOP_norm and val and val == SOP_norm:
                chosen_col = col_letter
                break
    else:  # 'Y'
        start_ci = 25  # Y
        ci = start_ci
        while ci <= max_c:
            val = _norm_str(ws.cell(row=2, column=ci).value)
            col_letter = get_column_letter(ci)
            print(f"[DEBUG] Checking {col_letter}2 = {val}")
            if val is None or val == "":
                # เจอช่องว่างครั้งแรก ให้หยุดช่วงของกลุ่ม Y
                break
            if SOP_norm and val == SOP_norm:
                chosen_col = col_letter
                break
            ci += 1

    if chosen_col is None:
        print("[DEBUG] row2 can't find matching SOP in the selected group range")
        return 0

    print(f"[DEBUG] Matched SOP at column {chosen_col}")

    # Scan rows from row 3 downward
    max_r = ws.max_row or 3
    yes_rows = []
    yes_blank_ag = 0
    no_rows = []
    no_nonblank_ag = 0

    for r in range(3, max_r + 1):
        val = _cell_value(ws, f"{chosen_col}{r}")
        ag = _cell_value(ws, f"AG{r}")
        if val is None:
            continue
        vu = val.upper()
        if vu == "YES":
            yes_rows.append(r)
            if ag is None or ag == "":
                yes_blank_ag += 1
        elif vu == "NO":
            no_rows.append(r)
            if ag is not None and ag != "":
                no_nonblank_ag += 1

    yes_all = "False" if yes_rows and (yes_blank_ag > 0) else "True"
    no_inapp = "False" if no_rows and (no_nonblank_ag > 0) else "True"

    print(f"[DEBUG] Total YES rows = {len(yes_rows)}, YES rows with AG blank = {yes_blank_ag}")
    print(f"[DEBUG] Total NO rows  = {len(no_rows)}, NO rows with AG non-blank = {no_nonblank_ag}")
    print(f"[DEBUG] => Yes_All_applicable_TC_Presented = {yes_all}")
    print(f"[DEBUG] => No_Inapplicable_TC_Presented    = {no_inapp}")

    return yes_all, no_inapp



def sw_info_tracking(path: str):
    """
    Normalize path แล้วดึงค่า SP, SOP, SW, CARLINE, HW, TEST_LEVEL ตามโครงสร้าง:
        .../Execution/{SP}/{SOP}/{SW}/{CARLINE}/{HW}/Tracking/{TEST_LEVEL}/...

    - รองรับทั้ง path ของไฟล์และโฟลเดอร์
    - ถ้าไม่ได้อยู่ตามโครงสร้างด้านบน จะพยายามหาแบบ heuristic จากชื่อโฟลเดอร์
    - พิมพ์ผลสำหรับ debug ใน terminal
    - คืนค่า dict สำหรับใช้งานต่อ (ถ้าต้องการ)
    """
    # 1) Normalize path
    norm = os.path.normpath(path)

    # ถ้าเป็นไฟล์ ให้ใช้โฟลเดอร์ที่ครอบมัน
    # ตรวจนามสกุลทั่วไปของ Excel หากตรงถือว่าเป็นไฟล์
    ext = os.path.splitext(norm)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        base_dir = os.path.dirname(norm)
    else:
        base_dir = norm  # เป็นโฟลเดอร์อยู่แล้ว หรือไม่ทราบแน่ชัดก็ถือเป็นโฟลเดอร์

    # แยกเป็นส่วนๆ ตามตัวคั่นของ OS
    parts = base_dir.split(os.sep)
    parts_lower = [p.lower() for p in parts]

    # 2) เตรียมตัวแปรผลลัพธ์
    SP = SOP = SW = CARLINE = HW = TEST_LEVEL = None

    # 3) พยายามอิงโครงแบบ canonical: .../Execution/SP/SOP/SW/CARLINE/HW/Tracking/TEST_LEVEL/...
    def idx_of_exact(name: str):
        try:
            return parts_lower.index(name.lower())
        except ValueError:
            return -1

    exec_idx = idx_of_exact("execution")
    tracking_idx = idx_of_exact("tracking")

    def safe_get(i):
        return parts[i] if 0 <= i < len(parts) else None

    if exec_idx != -1:
        candidate_SP = safe_get(exec_idx + 1)
        candidate_SOP = safe_get(exec_idx + 2)
        candidate_SW = safe_get(exec_idx + 3)
        candidate_CARLINE = safe_get(exec_idx + 4)
        candidate_HW = safe_get(exec_idx + 5)
        # ควรมี "Tracking" ที่ exec_idx + 6
        candidate_tracking = safe_get(exec_idx + 6)
        candidate_TEST_LEVEL = safe_get(exec_idx + 7)

        # Validate เบื้องต้นด้วย pattern
        if candidate_SP and candidate_SP.upper().startswith("SP"):
            SP = candidate_SP
        if candidate_SOP and candidate_SOP.upper().startswith("SOP"):
            SOP = candidate_SOP
        if candidate_SW:
            SW = candidate_SW
        if candidate_CARLINE:
            CARLINE = candidate_CARLINE
        if candidate_HW and candidate_HW.upper().startswith("HW"):
            HW = candidate_HW
        if (candidate_tracking and candidate_tracking.lower() == "tracking") and candidate_TEST_LEVEL:
            TEST_LEVEL = candidate_TEST_LEVEL

    # 4) ถ้าอย่างใดอย่างหนึ่งยังว่าง ลองหาแบบ heuristic (ค้นจากทุก segment)
    #    - SP: เริ่มด้วย SP ตามด้วยตัวเลข/ตัวอักษร
    #    - SOP: เริ่มด้วย SOP ตามด้วยตัวเลข/จุด
    #    - SW: ตัวเลขที่มีจุด เช่น 2535.0 (หรือเก็บตามชื่อโฟลเดอร์ถัดจาก SOP ถ้าเข้า pattern)
    #    - HW: เริ่มด้วย HW
    #    - TEST_LEVEL: โฟลเดอร์ถัดจาก "Tracking"
    if tracking_idx != -1 and TEST_LEVEL is None:
        TEST_LEVEL = safe_get(tracking_idx + 1)

    # หา SP
    if SP is None:
        for p in parts:
            if re.fullmatch(r"(?i)SP[\w.-]+", p):
                SP = p
                break

    # หา SOP
    if SOP is None:
        for i, p in enumerate(parts):
            if re.fullmatch(r"(?i)SOP[\w.-]+", p):
                SOP = p
                # เดาว่า SW อาจอยู่ถัดไปถ้ายังไม่มี
                if SW is None:
                    nxt = safe_get(i + 1)
                    if nxt:
                        SW = nxt
                break

    # หา HW
    if HW is None:
        for p in parts:
            if re.fullmatch(r"(?i)HW[\w.-]*", p):
                HW = p
                break

    # หา CARLINE (เดา: โฟลเดอร์ที่อยู่ระหว่าง SW กับ HW ถ้าเจอทั้งคู่)
    if CARLINE is None and SW and HW:
        try:
            i_sw = parts.index(SW)
            i_hw = parts.index(HW)
            if i_sw + 1 < i_hw:
                CARLINE = parts[i_sw + 1]
        except ValueError:
            pass

    # ถ้า SW ยังไม่เจอ ให้ลองหาโฟลเดอร์ที่เป็นรูปเลขมีจุด เช่น 2535.0
    if SW is None:
        for p in parts:
            if re.fullmatch(r"\d+(?:\.\d+)?", p):
                SW = p
                break

    # 5) Debug print (ตามที่ขอ ไม่โชว์บน GUI)
    print("\n[sw_info_tracking] ===========================")
    print(f"[sw_info_tracking] Normalized path : {norm}")
    print(f"[sw_info_tracking] Base directory  : {base_dir}")
    print(f"[sw_info_tracking] Parts            : {parts}")
    print(f"[sw_info_tracking] SP        = {SP}")
    print(f"[sw_info_tracking] SOP       = {SOP}")
    print(f"[sw_info_tracking] SW        = {SW}")
    print(f"[sw_info_tracking] CARLINE   = {CARLINE}")
    print(f"[sw_info_tracking] HW        = {HW}")
    print(f"[sw_info_tracking] TEST_LEVEL= {TEST_LEVEL}")
    print("[sw_info_tracking] ===========================\n")

    return {
        "SP": SP,
        "SOP": SOP,
        "SW": SW,
        "CARLINE": CARLINE,
        "HW": HW,
        "TEST_LEVEL": TEST_LEVEL,
        "normalized_path": norm,
        "base_dir": base_dir,
    }



def browse_file():
    path = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm")],
    )
    if not path:
        return
    file_label.configure(text=path)
    fullpath_var.set(path)

    # เก็บค่าจาก path
    global info_vars
    info_vars = sw_info_tracking(path)

    # เริ่มการวิเคราะห์แบบไม่บล็อก GUI
    start_analysis(path, info_vars)




def show_sheets_and_analysis(path: str, info: dict):
    """
    ฟังก์ชันเดิม (คงไว้สำหรับอ้างอิง) ปัจจุบันเปลี่ยนไปใช้แบบ async แทน
    """
    textbox.configure(state="normal")
    textbox.delete("0.0", "end")
    textbox.insert("end", "Processing in background...\n")
    textbox.configure(state="disabled")


def start_analysis(path: str, info: dict):
    """
    เริ่มงานวิเคราะห์ไฟล์แบบไม่บล็อก GUI โดยใช้ Thread + Queue
    ประมวลผลเฉพาะ sheet 5-22 และแสดงผลในรูปแบบ Results: Each pages และ Result: Test Level
    """
    global result_queue, analysis_thread, cancel_flag
    cancel_flag = False
    result_queue = queue.Queue()

    # เตรียม UI
    textbox.configure(state="normal")
    textbox.delete("0.0", "end")
    textbox.insert("end", "Processing...\n")
    textbox.configure(state="disabled")

    status_label.configure(text="Opening File...")
    progress_bar.configure(mode="indeterminate")
    progress_bar.start()
    browse_btn.configure(state="disabled")
    clear_btn.configure(state="disabled")

    def worker():
        try:
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            sheets = wb.sheetnames or []
            
            # ประมวลผลเฉพาะ sheet 5-22 (index 4-21)
            target_sheets = sheets[4:22] if len(sheets) >= 22 else sheets[4:]
            total = len(target_sheets)
            result_queue.put(("__meta__", {"total": total}))
            
            # เก็บผลลัพธ์ของแต่ละ sheet
            sheet_results = {}  # {sheet_index: (sheet_name, yes_all, no_inapp)}
            
            for idx, s in enumerate(target_sheets, start=5):
                try:
                    ws = wb[s]
                    res = Check_Yes_All_applicable_TC_Presented_And_No_Inapplicable_TC_Presented(ws, info.get('SP'), info.get('SOP'))
                    if res == 0:
                        sheet_results[idx] = (s, None, None)
                    else:
                        yes_all, no_inapp = res
                        sheet_results[idx] = (s, yes_all, no_inapp)
                    
                    # อัปเดตความคืบหน้า
                    current = idx - 4  # 5->1, 6->2, ..., 22->18
                    result_queue.put(("progress", {"current": current, "total": total}))
                except Exception as se:
                    sheet_results[idx] = (s, None, None)
                    current = idx - 4
                    result_queue.put(("progress", {"current": current, "total": total}))
            
            # ส่งผลลัพธ์ทั้งหมดไป
            result_queue.put(("results", sheet_results))
            
            # เรียกใช้ validate_testlog
            # สร้าง base_dir สำหรับ TestlogReader: .../Tracking/{TEST_LEVEL}
            base_dir_for_testlog = None
            if info.get('base_dir'):
                parts = info['base_dir'].split(os.sep)
                parts_lower = [p.lower() for p in parts]
                try:
                    tracking_idx = parts_lower.index('tracking')
                    if tracking_idx + 1 < len(parts):
                        # สร้าง path ไปยัง Tracking/{TEST_LEVEL}
                        test_level = parts[tracking_idx + 1]
                        base_dir_for_testlog = os.sep.join(parts[:tracking_idx + 2])
                    else:
                        # ถ้ายังไม่มี TEST_LEVEL ให้ใช้ directory ของ Tracking
                        base_dir_for_testlog = os.sep.join(parts[:tracking_idx + 1])
                except ValueError:
                    pass
            
            if base_dir_for_testlog:
                testlog_result = validate_testlog(base_dir_for_testlog)
                result_queue.put(("testlog", testlog_result))
            
        except Exception as e:
            result_queue.put(("error", str(e)))
        finally:
            result_queue.put(("done", None))

    analysis_thread = threading.Thread(target=worker, daemon=True)
    analysis_thread.start()
    root.after(50, poll_results)


def calculate_test_level_results(sheet_results):
    """
    คำนวณผล Test Level จาก sheet results
    Returns: dict {test_level: (yes_all, no_inapp)}
    """
    test_level_results = {}
    
    # L1_FI: sheet 5
    if 5 in sheet_results:
        name, yes_all, no_inapp = sheet_results[5]
        if yes_all is not None and no_inapp is not None:
            test_level_results["L1_FI"] = (yes_all, no_inapp)
    
    # L2_FuSa: sheet 6-11
    l2_results = []
    for idx in range(6, 12):
        if idx in sheet_results:
            name, yes_all, no_inapp = sheet_results[idx]
            if yes_all is not None and no_inapp is not None:
                l2_results.append((yes_all, no_inapp))
    if l2_results:
        yes_all_all = all(y == "True" for y, _ in l2_results)
        no_inapp_all = all(n == "True" for _, n in l2_results)
        test_level_results["L2_FuSa"] = ("True" if yes_all_all else "False", 
                                          "True" if no_inapp_all else "False")
    
    # L3_QM: sheet 13-16
    l3_results = []
    for idx in range(13, 17):
        if idx in sheet_results:
            name, yes_all, no_inapp = sheet_results[idx]
            if yes_all is not None and no_inapp is not None:
                l3_results.append((yes_all, no_inapp))
    if l3_results:
        yes_all_all = all(y == "True" for y, _ in l3_results)
        no_inapp_all = all(n == "True" for _, n in l3_results)
        test_level_results["L3_QM"] = ("True" if yes_all_all else "False", 
                                        "True" if no_inapp_all else "False")
    
    # HW_Delta1.6: sheet 18
    if 18 in sheet_results:
        name, yes_all, no_inapp = sheet_results[18]
        if yes_all is not None and no_inapp is not None:
            test_level_results["HW_Delta1.6"] = (yes_all, no_inapp)
    
    # HW_Delta2.0: sheet 19
    if 19 in sheet_results:
        name, yes_all, no_inapp = sheet_results[19]
        if yes_all is not None and no_inapp is not None:
            test_level_results["HW_Delta2.0"] = (yes_all, no_inapp)
    
    # HW_Delta3.0: sheet 20
    if 20 in sheet_results:
        name, yes_all, no_inapp = sheet_results[20]
        if yes_all is not None and no_inapp is not None:
            test_level_results["HW_Delta3.0"] = (yes_all, no_inapp)
    
    return test_level_results


def validate_testlog(base_dir: str):
    """
    หาไฟล์ชื่อ TestlogReader ใน base_dir
    อ่านไฟล์ใน column B-K ตั้งแต่ row 5 ลงไปจนกว่าจะเจอ blank
    แสดงค่าที่ไม่ซ้ำในแต่ละ column
    
    Returns: dict {column_name: [unique_values]} หรือ None ถ้าไม่พบไฟล์
    """
    if not base_dir or not os.path.isdir(base_dir):
        print(f"[validate_testlog] Base directory not found: {base_dir}")
        return None
    
    # หาไฟล์ TestlogReader (รองรับทั้ง .xlsx, .xlsm, .xltx, .xltm)
    testlog_file = None
    for ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        candidate = os.path.join(base_dir, f"TestlogReader{ext}")
        if os.path.isfile(candidate):
            testlog_file = candidate
            break
    
    if not testlog_file:
        print(f"[validate_testlog] TestlogReader file not found in: {base_dir}")
        return None
    
    print(f"[validate_testlog] Found TestlogReader: {testlog_file}")
    
    try:
        wb = load_workbook(filename=testlog_file, read_only=True, data_only=True)
        ws = wb.active  # ใช้ sheet แรก
        
        # Column mapping: B=2, C=3, ..., K=11
        column_mapping = {
            'B': 'SWFK',
            'C': 'HWEL',
            'D': 'PIC version',
            'E': 'HW version',
            'F': 'Setup',
            'G': 'Configuration',
            'H': 'CAN_Database',
            'I': 'LIN_Database',
            'J': 'PDX',
            'K': 'Remark'
        }
        
        # เก็บค่าที่ไม่ซ้ำสำหรับแต่ละ column
        column_values = {col: set() for col in column_mapping.keys()}
        
        # อ่านตั้งแต่ row 5 ลงไปจนเจอ blank
        max_r = ws.max_row or 5
        start_row = 5
        
        for row in range(start_row, max_r + 1):
            # ตรวจสอบว่ามีค่าในแถวนี้หรือไม่ (เช็ค column B เป็นหลัก)
            b_val = _cell_value(ws, f"B{row}")
            if b_val is None or b_val == "":
                # ถ้า column B ว่าง ให้หยุด (อาจจะเจอ blank แถวแล้ว)
                # แต่ยังต้องเช็ค column อื่นๆ ด้วยในแถวเดียวกัน
                all_blank = True
                for col in column_mapping.keys():
                    val = _cell_value(ws, f"{col}{row}")
                    if val is not None and val != "":
                        all_blank = False
                        column_values[col].add(val)
                if all_blank:
                    # ถ้าทุก column ว่าง ให้หยุดการอ่าน
                    break
            else:
                # ถ้า column B มีค่า ให้อ่านทุก column
                for col in column_mapping.keys():
                    val = _cell_value(ws, f"{col}{row}")
                    if val is not None and val != "":
                        column_values[col].add(val)
        
        # แปลง set เป็น sorted list เพื่อให้ผลลัพธ์สม่ำเสมอ
        result = {}
        for col, label in column_mapping.items():
            unique_vals = sorted(list(column_values[col]), key=lambda x: str(x).lower())
            result[label] = unique_vals
        
        wb.close()
        return result
        
    except Exception as e:
        print(f"[validate_testlog] Error reading file: {e}")
        return None


def poll_results():
    """
    ดึงผลจาก Queue และอัปเดต UI เป็นช่วงๆ เพื่อให้ GUI ลื่นไหล
    """
    try:
        while True:
            item_type, payload = result_queue.get_nowait()
            if item_type == "__meta__":
                total = payload.get("total", 0)
                if total > 0:
                    progress_bar.stop()
                    progress_bar.configure(mode="determinate")
                    progress_bar.set(0)
                    status_label.configure(text=f"found {total} Analysis...")
                    poll_results.total = total
                    poll_results.processed = 0
                else:
                    status_label.configure(text="Not Found any Analysis sheets.")
            elif item_type == "progress":
                current = payload.get("current", 0)
                total = payload.get("total", 0)
                if total > 0:
                    progress = current / total
                    progress_bar.set(progress)
                    status_label.configure(text=f"Analysis done {current}/{total} Sheets...")
            elif item_type == "results":
                # แสดงผลลัพธ์
                sheet_results = payload
                textbox.configure(state="normal")
                textbox.delete("0.0", "end")
                
                # ส่วนที่ 1: Results: Each pages
                textbox.insert("end", "Results: Each pages\n")
                textbox.insert("end", "----------------------------------------------\n")
                for idx in sorted(sheet_results.keys()):
                    name, yes_all, no_inapp = sheet_results[idx]
                    if yes_all is None or no_inapp is None:
                        textbox.insert("end", f"{idx}. {name} -> row1/row2 can't find matching -> 0\n")
                    else:
                        textbox.insert("end", 
                            f"{idx}. {name} -> Yes_All_applicable_TC_Presented = {yes_all}; "
                            f"No_Inapplicable_TC_Presented = {no_inapp}\n")
                
                textbox.insert("end", "----------------------------------------------\n")
                
                # ส่วนที่ 2: Result: Test Level
                test_level_results = calculate_test_level_results(sheet_results)
                textbox.insert("end", "Result: Test Level\n")
                for test_level in ["L1_FI", "L2_FuSa", "L3_QM", "HW_Delta1.6", "HW_Delta2.0", "HW_Delta3.0"]:
                    if test_level in test_level_results:
                        yes_all, no_inapp = test_level_results[test_level]
                        textbox.insert("end", 
                            f"{test_level} -> Yes_All_applicable_TC_Presented = {yes_all}; "
                            f"No_Inapplicable_TC_Presented = {no_inapp}\n")
                
                textbox.see("end")
                textbox.configure(state="disabled")
                
            elif item_type == "testlog":
                # แสดงผล Test_log Filter
                testlog_result = payload
                if testlog_result:
                    textbox.configure(state="normal")
                    
                    # หา position เพื่อ insert ต่อจาก Result: Test Level
                    textbox.insert("end", "------------------------------------------------\n")
                    textbox.insert("end", "Test_log Filter:\n")
                    
                    # Mapping ตาม requirement
                    column_order = [
                        'SWFK',
                        'HWEL',
                        'PIC version',
                        'HW version',
                        'Setup',
                        'Configuration',
                        'CAN_Database',
                        'LIN_Database',
                        'PDX',
                        'Remark'
                    ]
                    
                    for label in column_order:
                        if label in testlog_result and testlog_result[label]:
                            textbox.insert("end", f"{label}\n")
                            for val in testlog_result[label]:
                                textbox.insert("end", f"- {val}\n")
                    
                    textbox.see("end")
                    textbox.configure(state="disabled")
                
            elif item_type == "error":
                messagebox.showerror("Error", f"Failed to read workbook:\n{payload}")
            elif item_type == "done":
                status_label.configure(text="Analysis completed.")
                progress_bar.stop()
                browse_btn.configure(state="normal")
                clear_btn.configure(state="normal")
                return
    except queue.Empty:
        pass

    root.after(50, poll_results)



def clear_selection():
    fullpath_var.set("")
    file_label.configure(text="No file selected")
    textbox.configure(state="normal")
    textbox.delete("0.0", "end")
    textbox.configure(state="disabled")


def main():
    global root
    ctk.set_appearance_mode("system")
    ctk.set_default_color_theme("blue")

    root = ctk.CTk()
    root.title("Analysis Tracking Sheet")
    root.geometry("800x600")

    frame = ctk.CTkFrame(root, corner_radius=8)
    frame.pack(fill="both", expand=True, padx=16, pady=16)

    header = ctk.CTkLabel(frame, text="Analysis Tracking Sheet (Excel .xlsx)", font=ctk.CTkFont(size=16, weight="bold"))
    header.pack(pady=(8, 12))

    btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
    btn_frame.pack(fill="x", pady=(0, 8), padx=8)

    global browse_btn, clear_btn
    browse_btn = ctk.CTkButton(btn_frame, text="Browse...", width=120, command=browse_file)
    browse_btn.pack(side="left")

    clear_btn = ctk.CTkButton(btn_frame, text="Clear", width=80, command=clear_selection)
    clear_btn.pack(side="left", padx=(8, 0))

    global file_label, textbox, fullpath_var
    fullpath_var = tk.StringVar(value="")

    file_label = ctk.CTkLabel(frame, text="No file selected", anchor="w")
    file_label.pack(fill="x", padx=8, pady=(4, 8))

    # Progress and status
    global progress_bar, status_label
    progress_bar = ctk.CTkProgressBar(frame)
    progress_bar.pack(fill="x", padx=8, pady=(0, 4))
    progress_bar.set(0)

    status_label = ctk.CTkLabel(frame, text="Ready to process", anchor="w")
    status_label.pack(fill="x", padx=8, pady=(0, 8))

    # Textbox to show sheet names
    textbox = ctk.CTkTextbox(frame, width=520, height=200)
    textbox.pack(padx=8, pady=(0, 8), fill="both", expand=True)
    textbox.configure(state="disabled")

    root.mainloop()


if __name__ == "__main__":
    main()